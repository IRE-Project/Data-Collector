import json
from openpyxl import load_workbook
import os
from collections import defaultdict
import sys


def extract_data(path, json_path):
    data = defaultdict(lambda : {})
    count = 0
    sys.stderr.write(str(count) + "\n")

    for file in os.listdir(path):
        sys.stderr.write(f"\r{count}")
        count += 1

        try:
            wb = load_workbook(path+file)
            sheet_obj = wb.active

            CIN = file.split(".")[0]
            if CIN != sheet_obj.cell(2, 3).value:
                print("Integrity Error: File Integrity ", file)
            elif data[CIN] != {}:
                print("Integrity Error: Duplicate ", file)
            else:
                record = {}
                for row in range(2,19):
                    record[sheet_obj.cell(row, 1).value] = sheet_obj.cell(row, 3).value

                row = 22
                charges = []
                while sheet_obj.cell(row, 1) != None:
                    charges.append([sheet_obj.cell(row, col).value for col in range(1,6)])
                    row += 1
                record["Charges"] = charges

                row +=1
                directors = []
                while sheet_obj.cell(row, 1) != None:
                    directors.append([sheet_obj.cell(row, col).value for col in range(1, 5)])
                    row += 1
                record["Directors/Signatory Details"] = directors

                data[CIN] = record

        except Exception as e:
            print(str(e), file)

    file = open(json_path, "w+")
    json.dump(data, file, indent=4)
    file.close()


if __name__=="__main__":
    path = "../../data/mca.gov.in/mca_Data(125k)/V1/"
    json_path = "../../data/mca.gov.in/mca_Data(125k)/V1_extracted.json"
    extract_data(path, json_path)