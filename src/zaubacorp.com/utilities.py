import json
from openpyxl import load_workbook


def dictlinks_to_list():
    file = open("../../data/zaubacorp.com/links.json", "r")
    links = json.load(file)
    file.close()

    data = []
    for key, val in links.items():
        row = [key, val["Company"], val["RoC"], val["Status"], val["link"]]
        data.append(row)

    print(len(data))

    file = open("../../data/zaubacorp.com/linkslist.json", "w+")
    json.dump({"data": data}, file, indent=4)
    file.close()


def match_old_new():
    """
    Checks if all the companies in the old list is there in the new list
    """
    wb = load_workbook(filename="../../data/company_master_data_upto_Mar_2015_Telangana.xlsx")
    sheet_obj = wb.active

    file = open("../../data/zaubacorp.com/links.json", "r")
    links = json.load(file)
    file.close()

    cin_list = []

    count = 0
    for row in range(2, 77042):
        cin = sheet_obj.cell(row, 1).value
        try:
            x = links[cin]
        except KeyError as e:
            print(cin)
            count += 1
            cin_list.append(cin)

    print("Data Count Not in New Data: ", count)

    file = open("../../data/zaubacorp.com/extra_cin_in_master.json", "w+")
    json.dump({"CIN": cin_list}, file, indent=4)
    file.close()


match_old_new()
