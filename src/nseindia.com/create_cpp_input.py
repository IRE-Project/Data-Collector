"""@file
This file is responsible for creating a file that can act as input for similarity.cpp
"""
import json
from openpyxl import load_workbook
from collections import defaultdict

def excel_to_json(file_path):
    """
    Format the excel data to dict format
    :param file_path: path to excel
    :return: dict formatted excel data
    """
    wb = load_workbook(file_path)
    sheet_obj = wb.active

    data = defaultdict(lambda: {})
    row = 2
    while True:
        sno = sheet_obj.cell(row, 1).value
        if sno is None or sno == "":
            break
        symbol = sheet_obj.cell(row, 2).value
        c_name = sheet_obj.cell(row, 3).value
        capitalization = sheet_obj.cell(row, 4).value

        if data[c_name] == {}:
            data[c_name] = {
                "symbol": symbol,
                "capitalization": capitalization
            }
        else:
            print("Integrity Error: ", c_name)

        row += 1
    return data


def load_cdata(file_path):
    """
    Loads json data from file_path
    :param file_path: path to json file
    :return: json parsed data
    """
    file = open(file_path, 'r')
    data = json.load(file)
    file.close()
    return data


def create_cpp_input(nse, c_data, file_path):
    """
    Create the input file to act as input for similarity.cpp
    :param nse: the raw nse json data
    :param c_data: the links list exhaustive data
    :param file_path: the path to file that will act as input for similarity.cpp
    :return: None
    """
    file = open(file_path, "w+")
    file.write(str(len(nse)) + "\n")
    for key, val in nse.items():
        file.write(key.lower().replace(" ", "_") + "\n")
        if val["capitalization"] == "* Not Traded as on March 31, 2021":
            file.write("0\n")
        else:
            file.write(str(val["capitalization"]) + "\n")
        file.write(val["symbol"] + "\n")

    file.write(str(len(c_data)) + "\n")
    for key, val in c_data.items():
        file.write(key + "\n")
        file.write(val["Company"].lower().replace(" ", "_") + "\n")

    file.close()


if __name__ == '__main__':
    nse_json = excel_to_json("../../data/nseindia.com/MCAP31032021_0.xlsx")
    c_data = load_cdata('../../data/zaubacorp.com/links.json')
    create_cpp_input(nse_json, c_data, "inp.txt")
