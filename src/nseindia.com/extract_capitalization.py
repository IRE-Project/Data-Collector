import json
from openpyxl import load_workbook
from collections import defaultdict
import nltk

def excel_to_json(file_path):
    wb = load_workbook(file_path)
    sheet_obj = wb.active

    data = defaultdict(lambda: {})
    row = 2
    while True:
        sno = sheet_obj.cell(row, 1).value
        if sno is None or sno == "":
            break
        symbol = sheet_obj.cell(row, 2).value
        c_name = sheet_obj.cell(row, 3).value
        capitalization = sheet_obj.cell(row, 4).value

        if data[c_name] == {}:
            data[c_name] = {
                "symbol": symbol,
                "capitalization": capitalization
            }
        else:
            print("Integrity Error: ", c_name)

        row += 1
    return data


def load_cdata(file_path):
    file = open(file_path, 'r')
    data = json.load(file)
    file.close()
    return data


def compare_add(nse, company, save_path):

    extracted_data = {}
    progress = 0

    for cin in company.keys():
        c_name = company[cin]["Company"]
        extracted_data[cin] = []

        for name,val in nse.items():
            ed = nltk.edit_distance(c_name.lower(), name.lower())
            similarity = 100 * (1 - ed / len(c_name))
            if similarity > 60:
                candidate = val
                candidate["name"] = name
                extracted_data[cin].append(candidate)

        print(f"\rProgress: {progress}", end="")
        progress += 1

    file = open(save_path, "w+")
    json.dump(extracted_data, file, indent=4)
    file.close()


if __name__ == '__main__':
    nse_json = excel_to_json("../../data/nseindia.com/MCAP31032021_0.xlsx")
    c_data = load_cdata('../../data/zaubacorp.com/links.json')
    compare_add(nse_json, c_data, save_path = "../../data/nseindia.com/nse_data.json")
